# -*- coding: utf-8 -*-
"""fi_desk_lookup_import.py — "Dropdown Listeler.xlsx" → FI lookup tabloları.

OFİSTE koşulur (Spyder → Run/F5; argparse YOK, davranış KONFİG
sabitlerinden). 2026-08-05 karar değişikliği: lookup'lar artık UYGULAMADAN
da yönetilir (/fi-desk/admin — IS_ADMIN'li kullanıcılar; spec §5b). Bu
script İLK KURULUM / excel'den toplu yükleme içindir ve hedef tablolarında
TAM YENİLEME yapar — koşarsa Yönetim ekranından yapılmış banka/exporter
değişikliklerinin ÜZERİNE YAZAR. Form dropdown'ları bootstrap API'siyle bu
tablolardan okunur; koşu sonrası SAYFAYI YENİLEMEK yeterlidir.

Excel beklenen yapı (kolon adları esnek — boşluk/büyük-küçük normalize
edilir):

  Sheet "Group, Country, Region"  →  FI_LU_BANK (TAM YENİLEME)
      Bank | REGION | COUNTRY | MAIN GROUP
      - COUNTRY/REGION lender ülke/bölge otomatiği için; MAIN GROUP →
        GROUP_COMPANY. Group şirketinin ülke/bölgesi excelde yok →
        MAIN GROUP adı Bank kolonunda ayrıca satırsa ondan alınır,
        yoksa boş kalır.
      - Ayrıca uniq COUNTRY→REGION çiftleri FI_LU_LIST'in COUNTRY
        listesine yazılır (ATTR1=region; Lender Region otomatiği buradan).
      - SELF_BANK excelde yoksa satır olarak eklenir ve IS_SELF=1
        işaretlenir (Borrower varsayılanı).

  Sheet "Lehtar"  →  FI_LU_LIST (LEHTAR_TARGET_LISTS listeleri, TAM YENİLEME)
      LİSTEDEKİ İSİM | LEHTAR ADI
      - VALUE_CD = listedeki isim (dropdown'da görünen), LABEL = lehtar adı.
      - Hedef yalnız EXPORTER (2026-08-04 kararı: Importer bizim müşterimiz —
        lehtar listesinden DOLDURULMAZ, form EDW müşteri tablosunda typeahead
        arar; bkz. app.py FI_DESK_CUSTOMER_TABLE ve docs/FI_DESK_SPEC.md §8).
        Eski koşulardan kalan IMPORTER listesi varsa dokunulmaz; yalnız
        müşteri tablosu konfigüre edilmemişse arama fallback'i olarak okunur.

Örnek/seed verisi temizliği: FI_LU_BANK ve hedef listeler tam yenilendiği
için schema job'unun YER TUTUCU kayıtları (örnek bankalar, EXP_SAMPLE_1,
örnek COUNTRY satırları) otomatik silinir. CURRENCY / REPAYMENT /
BASE_RATE / ESG / BUSINESS_SEGMENT / COVERAGE_PROVIDER /
ADDITIONAL_COST_TYPE listeleri excelde olmadığından DOKUNULMAZ; IMPORTER
listesi de artık hedef olmadığından olduğu gibi kalır.
"""
from __future__ import annotations

import sys
import unicodedata
from datetime import datetime
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO_ROOT))

# ─────────────────────────────────────────────────────────────────────────
# KONFİG — Spyder'da args çalışmadığı için buradan düzenle (argparse yok).
# ─────────────────────────────────────────────────────────────────────────
EXCEL_PATH = r"C:\Users\A63837\Desktop\Dropdown Listeler.xlsx"
SCHEMA: str | None = None      # None → bağlantı kullanıcısının şeması
SELF_BANK = "QNB"              # Borrower varsayılanı (IS_SELF=1). Excelde bu
                               # adla satır yoksa eklenir.
BANK_SHEET = "Group, Country, Region"
LEHTAR_SHEET = "Lehtar"
#: Lehtar listesi hangi dropdown'lara yazılsın. IMPORTER artık hedef DEĞİL
#: (müşteri tablosu typeahead'i kullanılır) — geçici olarak geri almak
#: gerekirse ("EXPORTER", "IMPORTER") yap.
LEHTAR_TARGET_LISTS: tuple[str, ...] = ("EXPORTER",)
# ─────────────────────────────────────────────────────────────────────────


def _norm(s) -> str:
    """Kolon adı normalize: trim + Türkçe-duyarsız upper + iç boşluk sadeleşir."""
    s = unicodedata.normalize("NFKC", str(s or "")).strip()
    s = " ".join(s.split())
    return s.replace("İ", "I").replace("ı", "i").upper()


def _cell(v) -> str | None:
    s = str(v).strip() if v is not None else ""
    return s if s and s.lower() != "nan" else None


def _find_col(columns: dict[str, int], *cands: str) -> int:
    for c in cands:
        if _norm(c) in columns:
            return columns[_norm(c)]
    raise SystemExit(f"Kolon bulunamadı: {cands} — mevcut: {sorted(columns)}")


def parse_bank_sheet(rows: list[list]) -> tuple[list[dict], list[tuple[str, str]]]:
    """(banka satırları, uniq (country, region) çiftleri).

    rows[0] başlık satırıdır. Banka satırı: BANK_NM, COUNTRY, REGION,
    GROUP_COMPANY (+ grup şirketi Bank kolonunda ayrıca satırsa onun
    ülke/bölgesi GROUP_COMPANY_COUNTRY/REGION'a çekilir)."""
    header = {_norm(v): i for i, v in enumerate(rows[0]) if _cell(v)}
    i_bank = _find_col(header, "Bank", "Banka")
    i_region = _find_col(header, "REGION", "Bölge")
    i_country = _find_col(header, "COUNTRY", "Ülke")
    i_group = _find_col(header, "MAIN GROUP", "Group", "Grup")

    banks: dict[str, dict] = {}
    for r in rows[1:]:
        name = _cell(r[i_bank]) if len(r) > i_bank else None
        if not name:
            continue
        banks[name] = {
            "BANK_NM": name,
            "COUNTRY": _cell(r[i_country]) if len(r) > i_country else None,
            "REGION": _cell(r[i_region]) if len(r) > i_region else None,
            "GROUP_COMPANY": _cell(r[i_group]) if len(r) > i_group else None,
        }

    # Grup şirketi banka listesinde ayrıca satırsa ülke/bölgesini taşı
    by_norm = {_norm(k): v for k, v in banks.items()}
    for b in banks.values():
        grp = b.get("GROUP_COMPANY")
        src = by_norm.get(_norm(grp)) if grp else None
        b["GROUP_COMPANY_COUNTRY"] = src["COUNTRY"] if src else None
        b["GROUP_COMPANY_REGION"] = src["REGION"] if src else None

    countries: dict[str, str] = {}
    for b in banks.values():
        if b["COUNTRY"] and b["COUNTRY"] not in countries:
            countries[b["COUNTRY"]] = b["REGION"] or ""
    return list(banks.values()), sorted(countries.items())


def parse_lehtar_sheet(rows: list[list]) -> list[tuple[str, str]]:
    """(VALUE_CD=listedeki isim, LABEL=lehtar adı) çiftleri."""
    header = {_norm(v): i for i, v in enumerate(rows[0]) if _cell(v)}
    i_kisa = _find_col(header, "LİSTEDEKİ İSİM", "LISTEDEKI ISIM")
    i_ad = _find_col(header, "LEHTAR ADI", "LEHTAR")
    out: list[tuple[str, str]] = []
    seen = set()
    for r in rows[1:]:
        kisa = _cell(r[i_kisa]) if len(r) > i_kisa else None
        ad = _cell(r[i_ad]) if len(r) > i_ad else None
        if not kisa or kisa in seen:
            continue
        seen.add(kisa)
        out.append((kisa, ad or kisa))
    return out


def _read_sheet(path: str, sheet: str) -> list[list]:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True, read_only=True)
    if sheet not in wb.sheetnames:
        raise SystemExit(f"Sheet yok: {sheet!r} — mevcut: {wb.sheetnames}")
    return [list(row) for row in wb[sheet].iter_rows(values_only=True)]


def main() -> int:
    from DataClient import DataClient

    bank_rows, countries = parse_bank_sheet(_read_sheet(EXCEL_PATH, BANK_SHEET))
    lehtar_rows = parse_lehtar_sheet(_read_sheet(EXCEL_PATH, LEHTAR_SHEET))
    print(f"Excel okundu: {len(bank_rows)} banka, {len(countries)} ülke, "
          f"{len(lehtar_rows)} lehtar")

    # SELF_BANK garanti + işaret
    if not any(_norm(b["BANK_NM"]) == _norm(SELF_BANK) for b in bank_rows):
        bank_rows.insert(0, {"BANK_NM": SELF_BANK, "COUNTRY": None,
                             "REGION": None, "GROUP_COMPANY": None,
                             "GROUP_COMPANY_COUNTRY": None,
                             "GROUP_COMPANY_REGION": None})
        print(f"   + {SELF_BANK} excelde yoktu, eklendi (IS_SELF=1)")

    dc = DataClient()
    con = dc.get_connection()
    try:
        schema = (SCHEMA or con.username).upper()
        now = datetime.now()
        cur = con.cursor()
        try:
            # ── FI_LU_BANK: tam yenileme (seed örnek bankaları da gider) ──
            cur.execute(f"DELETE FROM {schema}.FI_LU_BANK")
            cur.executemany(
                f"INSERT INTO {schema}.FI_LU_BANK "
                "(BANK_NM, COUNTRY, REGION, GROUP_COMPANY, "
                "GROUP_COMPANY_COUNTRY, GROUP_COMPANY_REGION, IS_SELF, "
                "ACTIVE_FLG, LOADED_AT) VALUES (:1, :2, :3, :4, :5, :6, :7, 1, :8)",
                [(b["BANK_NM"], b["COUNTRY"], b["REGION"], b["GROUP_COMPANY"],
                  b["GROUP_COMPANY_COUNTRY"], b["GROUP_COMPANY_REGION"],
                  1 if _norm(b["BANK_NM"]) == _norm(SELF_BANK) else 0, now)
                 for b in bank_rows])

            # ── COUNTRY listesi: tam yenileme (ATTR1 = region) ──
            cur.execute(f"DELETE FROM {schema}.FI_LU_LIST "
                        "WHERE LIST_NAME = 'COUNTRY'")
            cur.executemany(
                f"INSERT INTO {schema}.FI_LU_LIST "
                "(LIST_NAME, VALUE_CD, LABEL, PARENT_CD, ATTR1, SORT_NO, "
                "ACTIVE_FLG, LOADED_AT) "
                "VALUES ('COUNTRY', :1, :2, NULL, :3, :4, 1, :5)",
                [(c, c, r or None, i, now)
                 for i, (c, r) in enumerate(countries)])

            # ── Lehtar → hedef listeler: tam yenileme (sample'lar gider) ──
            for list_name in LEHTAR_TARGET_LISTS:
                cur.execute(f"DELETE FROM {schema}.FI_LU_LIST "
                            "WHERE LIST_NAME = :ln", {"ln": list_name})
                cur.executemany(
                    f"INSERT INTO {schema}.FI_LU_LIST "
                    "(LIST_NAME, VALUE_CD, LABEL, PARENT_CD, ATTR1, SORT_NO, "
                    f"ACTIVE_FLG, LOADED_AT) "
                    f"VALUES ('{list_name}', :1, :2, NULL, NULL, :3, 1, :4)",
                    [(cd, lbl, i, now)
                     for i, (cd, lbl) in enumerate(lehtar_rows)])

            con.commit()
        except Exception:
            con.rollback()
            raise
        finally:
            cur.close()

        print(f"✓ {schema}.FI_LU_BANK: {len(bank_rows)} banka "
              f"(IS_SELF: {SELF_BANK})")
        print(f"✓ {schema}.FI_LU_LIST COUNTRY: {len(countries)} ülke")
        for ln in LEHTAR_TARGET_LISTS:
            print(f"✓ {schema}.FI_LU_LIST {ln}: {len(lehtar_rows)} kayıt")
        print("Bitti — form dropdown'ları sayfa yenilenince yeni listelerle "
              "dolar (bootstrap API bu tablolardan okur).")
        return 0
    finally:
        dc.drop_connection(con)


if __name__ == "__main__":
    raise SystemExit(main())
